"""Upload endpoint for the Legal-maintained FATF baseline workbook.

Why this exists: the daily pipeline in strangeromo-cloud/aml reads the baseline from
seeds/fatf-baseline.xlsx in that repo, and Legal cannot reasonably be expected to
commit to GitHub. This accepts the workbook they edit, validates it, shows them what
would change, and — only after they confirm — commits it through the GitHub API. The
pipeline itself needs no change: its next run picks the new file up and the emailed
attachment carries it.

Two-step on purpose. This file is the authority the whole comparison rests on, so a
mis-uploaded workbook must never replace it silently: /validate parses and returns a
diff, /commit writes only what /validate already approved.

Env:
  BASELINE_UPLOAD_TOKEN  Shared secret required to upload; unset means uploads are off
  GH_REPO_TOKEN   GitHub token with contents:write on the pipeline repo
  GH_REPO         owner/name, default strangeromo-cloud/aml
  GH_BRANCH       default main
"""
from __future__ import annotations

import base64
import hmac
import json
import logging
import os
import re
import unicodedata
from datetime import datetime, timedelta, timezone
from io import BytesIO
from typing import Any, Optional

import httpx
from fastapi import APIRouter, File, Form, Header, HTTPException, UploadFile
from openpyxl import load_workbook
from pydantic import BaseModel

logger = logging.getLogger("aml-chat.fatf-baseline")
router = APIRouter(prefix="/api/aml/fatf-baseline", tags=["fatf-baseline"])

# The baseline is the authority the whole FATF comparison rests on, so the write path
# is gated. Reads are not: they only expose the FATF lists, which are public.
UPLOAD_TOKEN = os.getenv("BASELINE_UPLOAD_TOKEN", "").strip()

GH_REPO_TOKEN = os.getenv("GH_REPO_TOKEN", "").strip()
GH_REPO = os.getenv("GH_REPO", "strangeromo-cloud/aml").strip()
GH_BRANCH = os.getenv("GH_BRANCH", "main").strip()
BASELINE_PATH = "scripts/data-refresh/seeds/fatf-baseline.xlsx"
# Legal-confirmed overrides for the two auto-fetched lists. Written only when the
# uploaded workbook carries that sheet, so an upload cannot blank a list by omission.
CPI_OVERRIDE_PATH = "scripts/data-refresh/seeds/cpi-override.json"
OFFSHORE_OVERRIDE_PATH = "scripts/data-refresh/seeds/offshore-override.json"
SHEET_NAME = "FATF 黑灰名单"
TZ_SHANGHAI = timezone(timedelta(hours=8))

# Same bounds the pipeline enforces, so a workbook accepted here cannot be rejected
# there. Keep in step with verify_fatf.validate_rows.
BLACK_RANGE = (1, 6)
GREY_RANGE = (8, 40)
MAX_BYTES = 2 * 1024 * 1024

ALIASES = {
    "democratic peoples republic of korea": "north korea",
    "democratic republic of korea": "north korea",
    "korea democratic peoples republic of": "north korea",
    "dprk": "north korea",
    "burma": "myanmar",
    "virgin islands uk": "british virgin islands",
    "virgin islands british": "british virgin islands",
    "democratic republic of congo": "democratic republic of the congo",
    "lao peoples democratic republic": "laos",
    "lao pdr": "laos",
    "syrian arab republic": "syria",
    "united republic of tanzania": "tanzania",
    "viet nam": "vietnam",
    # TI writes the Koreas inverted; the FATF/Legal workbooks do not.
    "korea north": "north korea",
    "korea south": "south korea",
    "republic of korea": "south korea",
}


def require_token(supplied: Optional[str]) -> None:
    """Fail closed: an unconfigured token means writes are refused, never waved through.

    Compared with compare_digest so a wrong guess cannot be narrowed down by timing.
    """
    if not UPLOAD_TOKEN:
        raise HTTPException(503, "服务器未配置 BASELINE_UPLOAD_TOKEN，基线上传已停用。"
                                 "请在部署环境变量里设置后重试。")
    if not supplied or not hmac.compare_digest(supplied.strip(), UPLOAD_TOKEN):
        raise HTTPException(401, "上传口令不正确。")


def norm(name: str) -> str:
    decomposed = unicodedata.normalize("NFKD", name or "")
    ascii_only = "".join(c for c in decomposed if not unicodedata.combining(c))
    # Hyphens become spaces before non-letters are stripped, otherwise "Guinea-Bissau"
    # collapses to "guineabissau" while "Guinea Bissau" keeps its space and the two
    # spellings of one country read as two different countries.
    ascii_only = re.sub(r"[-–—/]", " ", ascii_only)
    s = re.sub(r"[^a-z\s]", "", ascii_only.lower()).strip()
    s = re.sub(r"\s+", " ", s)
    s = re.sub(r"^the ", "", s)
    return ALIASES.get(s, s)


def parse_workbook(data: bytes) -> dict[str, Any]:
    """Parse and validate an uploaded baseline workbook.

    Raises HTTPException(422) with a message Legal can act on — every rejection names
    what is wrong and where, because "上传失败" alone would just get retried.
    """
    if not data:
        raise HTTPException(422, "文件为空")
    if len(data) > MAX_BYTES:
        raise HTTPException(422, f"文件超过 {MAX_BYTES // 1024 // 1024} MB")
    try:
        wb = load_workbook(BytesIO(data), data_only=True)
    except Exception as e:
        raise HTTPException(422, f"无法作为 Excel 打开（{type(e).__name__}）。"
                                 f"请确认是 .xlsx 文件，不是 .xls 或 CSV。")
    if SHEET_NAME not in wb.sheetnames:
        raise HTTPException(422, f"找不到工作表「{SHEET_NAME}」。"
                                 f"当前工作表：{', '.join(wb.sheetnames)}")
    ws = wb[SHEET_NAME]

    note = str(ws.cell(1, 1).value or "")
    # The note often carries more than one date (a sheet exported from the daily
    # attachment ends with "抓取: <日期>"), so read the one behind the 名单日期 label
    # and only fall back to a bare date when the note has exactly one.
    labelled = re.search(r"(?:名单日期|名單日期|list\s*date)\s*[:：]?\s*(\d{4}-\d{2}-\d{2})",
                         note, re.I)
    bare = sorted(set(re.findall(r"\d{4}-\d{2}-\d{2}", note)))
    if labelled:
        list_date = labelled.group(1)
    elif len(bare) == 1:
        list_date = bare[0]
    elif len(bare) > 1:
        raise HTTPException(422, f"第 1 行有多个日期（{'、'.join(bare)}），无法判断哪个是名单日期。"
                                 f"请写成「名单日期: YYYY-MM-DD」。")
    else:
        raise HTTPException(422, "第 1 行说明里找不到名单日期。"
                                 "请在第 1 行写入 FATF 声明日期，格式「名单日期: YYYY-MM-DD」"
                                 "（例如 名单日期: 2026-10-23）。")

    black: list[str] = []
    grey: list[str] = []
    for r in range(3, ws.max_row + 1):
        lst, juris = ws.cell(r, 1).value, ws.cell(r, 2).value
        if not juris:
            continue
        label = str(lst or "")
        if "Black" in label or "黑" in label:
            black.append(str(juris).strip())
        elif "Grey" in label or "灰" in label:
            grey.append(str(juris).strip())
        else:
            raise HTTPException(422, f"第 {r} 行的「名单 List」既不是黑名单也不是灰名单："
                                     f"{label!r}。请填「黑名单 Black」或「灰名单 Grey」。")

    nb, ng = {norm(x) for x in black} - {""}, {norm(x) for x in grey} - {""}
    if not (BLACK_RANGE[0] <= len(nb) <= BLACK_RANGE[1]):
        raise HTTPException(422, f"黑名单 {len(nb)} 条，超出合理范围 {BLACK_RANGE[0]}–{BLACK_RANGE[1]} 条。"
                                 f"FATF 黑名单历来只有 1–5 个辖区，请检查是否填错列。")
    if not (GREY_RANGE[0] <= len(ng) <= GREY_RANGE[1]):
        raise HTTPException(422, f"灰名单 {len(ng)} 条，超出合理范围 {GREY_RANGE[0]}–{GREY_RANGE[1]} 条。")
    overlap = nb & ng
    if overlap:
        raise HTTPException(422, f"同一辖区同时出现在黑灰名单：{', '.join(sorted(overlap))}")

    return {"listDate": list_date, "black": sorted(black), "grey": sorted(grey),
            "normBlack": nb, "normGrey": ng, "rowCount": len(black) + len(grey)}


# ── The other two lists ─────────────────────────────────────────────────
# CPI and the offshore list are fetched automatically and that fetch is reliable, so a
# Legal upload must NOT overwrite them — only report where the two disagree. That is a
# question for a human ("whose CPI edition is this?"), not something to silently adopt.
# FATF is different: its page is unreachable, so there the uploaded list IS authority.
CPI_SNAPSHOT = "public/downloads/_snapshots/ti-cpi.json"
OFFSHORE_SNAPSHOT = "public/downloads/_snapshots/eu-offshore-centres.json"


def parse_cpi_sheet(wb) -> Optional[dict[str, Any]]:
    """{name: score} from whichever sheet is called "CPI <year>"."""
    title = next((t for t in wb.sheetnames if t.strip().upper().startswith("CPI")), None)
    if not title:
        return None
    ws = wb[title]
    scores: dict[str, Any] = {}
    for r in range(3, ws.max_row + 1):
        country, score = ws.cell(r, 2).value, ws.cell(r, 3).value
        if not country:
            continue
        try:
            scores[norm(str(country))] = int(round(float(score)))
        except (TypeError, ValueError):
            scores[norm(str(country))] = None
    m = re.search(r"(\d{4})", title)
    return {"sheet": title, "edition": m.group(1) if m else None, "scores": scores}


def parse_offshore_sheet(wb) -> Optional[dict[str, Any]]:
    title = next((t for t in wb.sheetnames if "Offshore" in t or "离岸" in t), None)
    if not title:
        return None
    ws = wb[title]
    names = []
    for r in range(3, ws.max_row + 1):
        v = ws.cell(r, 2).value
        if v:
            names.append(str(v).strip())
    return {"sheet": title, "names": names, "norm": {norm(n) for n in names} - {""}}


async def _snapshot(path: str) -> Optional[list[dict[str, Any]]]:
    r = await _gh("GET", f"/repos/{GH_REPO}/contents/{path}", params={"ref": GH_BRANCH})
    if r.status_code == 404:
        return None
    if r.status_code >= 400:
        raise HTTPException(502, f"读取 {path} 失败：{r.status_code}")
    return json.loads(base64.b64decode(r.json()["content"]))


def diff_cpi(fetched: list[dict[str, Any]], up: dict[str, Any]) -> dict[str, Any]:
    f = {norm(str(r.get("country"))): int(round(float(r["score"])))
         for r in fetched if r.get("score") is not None}
    u = up["scores"]
    only_up = sorted(set(u) - set(f))
    only_fetched = sorted(set(f) - set(u))
    changed = sorted(k for k in set(u) & set(f) if u[k] != f[k])
    return {
        "sheet": up["sheet"], "uploadEdition": up["edition"],
        "uploadCount": len(u), "fetchedCount": len(f),
        "onlyInUpload": only_up, "onlyInFetched": only_fetched,
        "scoreDiffers": [{"country": k, "upload": u[k], "fetched": f[k]} for k in changed[:60]],
        "scoreDiffersCount": len(changed),
        "identical": not (only_up or only_fetched or changed),
    }


def diff_offshore(fetched: list[dict[str, Any]], up: dict[str, Any]) -> dict[str, Any]:
    f = {norm(str(r.get("jurisdiction"))) for r in fetched} - {""}
    by_name = {norm(n): n for n in up["names"]}
    by_fetch = {norm(str(r.get("jurisdiction"))): str(r.get("jurisdiction")) for r in fetched}

    def nm(keys):
        return sorted(by_name.get(k) or by_fetch.get(k) or k for k in keys)

    only_up, only_f = up["norm"] - f, f - up["norm"]
    return {"sheet": up["sheet"], "uploadCount": len(up["norm"]), "fetchedCount": len(f),
            "onlyInUpload": nm(only_up), "onlyInFetched": nm(only_f),
            "identical": not (only_up or only_f)}


# ── GitHub ──────────────────────────────────────────────────────────────

async def _gh(method: str, path: str, **kw) -> httpx.Response:
    if not GH_REPO_TOKEN:
        # Say which operation is blocked: the same token gates reads and writes, and
        # "无法写入" on a history page just looks like a bug.
        what = "读取" if method.upper() == "GET" else "写入"
        raise HTTPException(503, f"服务器未配置 GH_REPO_TOKEN，无法{what}仓库。"
                                 f"请在部署环境变量里设置后重试。")
    headers = {"Authorization": f"Bearer {GH_REPO_TOKEN}",
               "Accept": "application/vnd.github+json",
               "X-GitHub-Api-Version": "2022-11-28"}
    async with httpx.AsyncClient(timeout=30) as c:
        return await c.request(method, f"https://api.github.com{path}", headers=headers, **kw)


async def fetch_current() -> tuple[bytes | None, str | None]:
    """(file bytes, blob sha) of the baseline currently in the repo."""
    r = await _gh("GET", f"/repos/{GH_REPO}/contents/{BASELINE_PATH}",
                  params={"ref": GH_BRANCH})
    if r.status_code == 404:
        return None, None
    if r.status_code >= 400:
        raise HTTPException(502, f"读取仓库现有基线失败：{r.status_code} {r.text[:180]}")
    j = r.json()
    return base64.b64decode(j["content"]), j["sha"]


async def _commit_files(files: dict[str, bytes], message: str) -> str:
    """Commit several files as ONE commit, via the Git data API.

    One submission has to be one version. Writing each file with the contents API
    produces a separate commit per file, which makes the history incoherent: the list
    is filtered by the baseline path, so a commit touching only the CPI override never
    appears, and the baseline commit's parent does not yet contain its sibling changes,
    so "what changed in this version" comes out wrong.

    Returns the new commit's html_url.
    """
    r = await _gh("GET", f"/repos/{GH_REPO}/git/ref/heads/{GH_BRANCH}")
    if r.status_code >= 400:
        raise HTTPException(502, f"读取分支 {GH_BRANCH} 失败：{r.status_code}")
    parent = r.json()["object"]["sha"]

    r = await _gh("GET", f"/repos/{GH_REPO}/git/commits/{parent}")
    if r.status_code >= 400:
        raise HTTPException(502, f"读取父提交失败：{r.status_code}")
    base_tree = r.json()["tree"]["sha"]

    tree_entries = []
    for path, content in files.items():
        rb = await _gh("POST", f"/repos/{GH_REPO}/git/blobs",
                       json={"content": base64.b64encode(content).decode(),
                             "encoding": "base64"})
        if rb.status_code >= 400:
            raise HTTPException(502, f"上传 {path} 内容失败：{rb.status_code} {rb.text[:140]}")
        tree_entries.append({"path": path, "mode": "100644", "type": "blob",
                             "sha": rb.json()["sha"]})

    rt = await _gh("POST", f"/repos/{GH_REPO}/git/trees",
                   json={"base_tree": base_tree, "tree": tree_entries})
    if rt.status_code >= 400:
        raise HTTPException(502, f"创建 tree 失败：{rt.status_code} {rt.text[:140]}")

    rc = await _gh("POST", f"/repos/{GH_REPO}/git/commits",
                   json={"message": message, "tree": rt.json()["sha"], "parents": [parent]})
    if rc.status_code >= 400:
        raise HTTPException(502, f"创建提交失败：{rc.status_code} {rc.text[:140]}")
    new_sha = rc.json()["sha"]

    rr = await _gh("PATCH", f"/repos/{GH_REPO}/git/refs/heads/{GH_BRANCH}",
                   json={"sha": new_sha})
    if rr.status_code >= 400:
        raise HTTPException(502, f"更新分支失败：{rr.status_code} {rr.text[:140]}")
    return rc.json().get("html_url") or f"https://github.com/{GH_REPO}/commit/{new_sha}"


def diff(cur: dict[str, Any] | None, new: dict[str, Any]) -> dict[str, Any]:
    if not cur:
        return {"firstUpload": True, "dateFrom": None, "dateTo": new["listDate"],
                "blackAdded": new["black"], "blackRemoved": [],
                "greyAdded": new["grey"], "greyRemoved": [], "identical": False}
    by_norm_new = {norm(x): x for x in new["black"] + new["grey"]}
    by_norm_cur = {norm(x): x for x in cur["black"] + cur["grey"]}

    def names(keys):
        return sorted(by_norm_new.get(k) or by_norm_cur.get(k) or k for k in keys)

    d = {
        "firstUpload": False,
        "dateFrom": cur["listDate"], "dateTo": new["listDate"],
        "blackAdded": names(new["normBlack"] - cur["normBlack"]),
        "blackRemoved": names(cur["normBlack"] - new["normBlack"]),
        "greyAdded": names(new["normGrey"] - cur["normGrey"]),
        "greyRemoved": names(cur["normGrey"] - new["normGrey"]),
    }
    d["identical"] = (cur["listDate"] == new["listDate"]
                      and not any(d[k] for k in
                                  ("blackAdded", "blackRemoved", "greyAdded", "greyRemoved")))
    return d


class CommitResp(BaseModel):
    committed: bool
    # Optional[...] rather than `str | None`: Pydantic evaluates model annotations at
    # class creation, and the PEP 604 form fails on Python 3.9 even under
    # `from __future__ import annotations`.
    commitUrl: Optional[str] = None
    message: str


@router.get("/status")
async def status() -> dict[str, Any]:
    """Is baseline upload actually usable right now?

    The token lives here, not in CI, so CI cannot inspect it — it asks this instead.
    Reports configuration, repo reachability, whether the token can actually write,
    and when it expires: a fine-grained PAT expires silently, and "discovered on the
    day Legal needs to update the baseline" is the failure mode worth avoiding.
    Returns no secret values.
    """
    out: dict[str, Any] = {
        # Which behaviours this deployment actually has. Without it, "is the feature
        # live or am I looking at a cached page?" is unanswerable from outside.
        "features": ["validate", "commit", "history", "status",
                     "compare-cpi", "compare-offshore", "overrides", "recency-precedence"],
        "uploadTokenConfigured": bool(UPLOAD_TOKEN),
        "repoTokenConfigured": bool(GH_REPO_TOKEN),
        "repo": f"{GH_REPO}:{GH_BRANCH}", "path": BASELINE_PATH,
        "repoReadable": False, "canWrite": None,
        "tokenExpiresAt": None, "baselineFound": None, "error": None,
    }
    if not GH_REPO_TOKEN:
        out["error"] = "GH_REPO_TOKEN 未配置"
        out["ready"] = False
        return out
    try:
        r = await _gh("GET", f"/repos/{GH_REPO}")
    except HTTPException as e:
        out["error"] = str(e.detail)
        out["ready"] = False
        return out
    # GitHub returns this header for fine-grained tokens; classic tokens omit it.
    out["tokenExpiresAt"] = r.headers.get("github-authentication-token-expiration")
    if r.status_code == 401:
        out["error"] = "GH_REPO_TOKEN 无效或已过期（GitHub 返回 401）"
        out["ready"] = False
        return out
    if r.status_code >= 400:
        out["error"] = f"读取仓库信息失败：{r.status_code}"
        out["ready"] = False
        return out
    out["repoReadable"] = True
    # permissions.push is the closest thing to "can write contents" without writing.
    out["canWrite"] = bool((r.json().get("permissions") or {}).get("push"))
    try:
        data, _ = await fetch_current()
        out["baselineFound"] = data is not None
        if data:
            out["listDate"] = parse_workbook(data)["listDate"]
    except HTTPException as e:
        out["error"] = str(e.detail)
    out["ready"] = bool(out["uploadTokenConfigured"] and out["repoReadable"]
                        and out["canWrite"] and out["baselineFound"])
    return out


@router.get("/auth")
async def auth(x_upload_token: Optional[str] = Header(None)) -> dict[str, Any]:
    """Lets the page check a token before the user picks a file."""
    require_token(x_upload_token)
    return {"ok": True}


@router.post("/validate")
async def validate(file: UploadFile = File(...),
                   x_upload_token: Optional[str] = Header(None)) -> dict[str, Any]:
    """Parse + validate the upload and show what would change. Writes nothing."""
    require_token(x_upload_token)
    data = await file.read()
    new = parse_workbook(data)
    # The other two sheets are optional: a workbook holding only the FATF sheet is a
    # perfectly good baseline upload.
    others: dict[str, Any] = {}
    try:
        wb = load_workbook(BytesIO(data), data_only=True)
        cpi_up, off_up = parse_cpi_sheet(wb), parse_offshore_sheet(wb)
        if cpi_up:
            snap = await _snapshot(CPI_SNAPSHOT)
            others["cpi"] = (diff_cpi(snap, cpi_up) if snap
                             else {"sheet": cpi_up["sheet"], "error": "仓库里没有 CPI 抓取快照"})
        if off_up:
            snap = await _snapshot(OFFSHORE_SNAPSHOT)
            others["offshore"] = (diff_offshore(snap, off_up) if snap
                                  else {"sheet": off_up["sheet"], "error": "仓库里没有离岸抓取快照"})
    except HTTPException as e:
        others["error"] = str(e.detail)
    except Exception as e:
        others["error"] = f"{type(e).__name__}: {e}"
    try:
        cur_bytes, _ = await fetch_current()
        cur = parse_workbook(cur_bytes) if cur_bytes else None
        repo_readable = True
        repo_error = None
    except HTTPException as e:
        # A diff is nice to have; a validation result is not worth losing because the
        # repo is briefly unreachable.
        cur, repo_readable, repo_error = None, False, e.detail
    return {
        "ok": True,
        "filename": file.filename,
        "listDate": new["listDate"],
        "counts": {"black": len(new["normBlack"]), "grey": len(new["normGrey"])},
        "black": new["black"], "grey": new["grey"],
        "diff": diff(cur, new) if repo_readable else None,
        "others": others,
        "repoReadable": repo_readable, "repoError": repo_error,
        "note": ("确认后提交，将写入仓库 "
                 f"{GH_REPO}:{GH_BRANCH}/{BASELINE_PATH}，"
                 "次日自动邮件的附件即使用新基线。"),
    }


@router.post("/commit", response_model=CommitResp)
async def commit(file: UploadFile = File(...),
                 uploader: str = Form(...),
                 confirmedDate: str = Form(...),
                 x_upload_token: Optional[str] = Header(None)) -> CommitResp:
    """Commit the workbook, but only if it still matches what was validated.

    confirmedDate is the list date the uploader saw in the preview: if the file has
    been swapped between preview and confirm, the dates disagree and this refuses.
    """
    require_token(x_upload_token)
    data = await file.read()
    new = parse_workbook(data)
    if new["listDate"] != confirmedDate.strip():
        raise HTTPException(409, f"文件与预览时不一致（预览名单日期 {confirmedDate}，"
                                 f"当前文件 {new['listDate']}）。请重新预览后再提交。")
    if not uploader.strip():
        raise HTTPException(422, "请填写上传人，提交记录需要留痕。")

    stamp = datetime.now(TZ_SHANGHAI).strftime("%Y-%m-%d %H:%M")
    files: dict[str, bytes] = {BASELINE_PATH: data}
    extras: list[str] = []
    try:
        wb = load_workbook(BytesIO(data), data_only=True)
        cpi_up, off_up = parse_cpi_sheet(wb), parse_offshore_sheet(wb)
    except Exception as e:
        raise HTTPException(422, f"解析工作簿失败：{type(e).__name__}: {e}")
    if cpi_up:
        files[CPI_OVERRIDE_PATH] = json.dumps({
            "confirmedBy": uploader.strip(), "confirmedAt": stamp,
            "edition": cpi_up["edition"], "sheet": cpi_up["sheet"],
            "scores": {k: v for k, v in cpi_up["scores"].items() if v is not None},
        }, ensure_ascii=False, indent=1, sort_keys=True).encode()
        extras.append(f"CPI {len(cpi_up['scores'])} 条")
    if off_up:
        files[OFFSHORE_OVERRIDE_PATH] = json.dumps({
            "confirmedBy": uploader.strip(), "confirmedAt": stamp,
            "jurisdictions": off_up["names"],
        }, ensure_ascii=False, indent=1, sort_keys=True).encode()
        extras.append(f"离岸 {len(off_up['names'])} 条")

    parts = [f"FATF 黑 {len(new['normBlack'])} / 灰 {len(new['normGrey'])}"] + extras
    url = await _commit_files(files, (
        f"chore(lists): 基线更新至 {new['listDate']}（{'；'.join(parts)}）\n\n"
        f"由 {uploader.strip()} 于 {stamp} 通过名单维护页提交。\n"
        f"次日 07:00（北京）定时运行的邮件附件将使用这些内容。"))
    logger.info("lists committed by %s → %s (%s)", uploader, new["listDate"], url)
    tail = ("　同时以法务版本为准：" + "、".join(extras)) if extras else ""
    return CommitResp(committed=True, commitUrl=url,
                      message=(f"已提交。基线名单日期 {new['listDate']}，"
                               f"黑 {len(new['normBlack'])} / 灰 {len(new['normGrey'])}。"
                               f"{tail}　次日 07:00（北京）自动邮件的附件将使用这些内容。"))


# ── Version history ─────────────────────────────────────────────────────
# The repo IS the history: every commit touching the baseline is a version, and git
# makes that record tamper-evident for free. The list view is built from commit
# messages so it costs one API call; the detail view fetches and parses that specific
# version on demand, which is what "查看当时的记录" needs.

_MSG_DATE = re.compile(r"基线更新至\s*(\d{4}-\d{2}-\d{2})")
_MSG_COUNTS = re.compile(r"黑\s*(\d+)\s*/\s*灰\s*(\d+)")
_MSG_UPLOADER = re.compile(r"由\s*(.+?)\s*于\s*([\d\-: ]+)\s*通过(?:基线上传页|名单维护页)提交")


def _from_message(msg: str) -> dict[str, Any]:
    """Metadata the upload page wrote into the commit message, when present.

    Commits made directly with git (not through this page) simply have none of it;
    the detail view still works because it parses the file itself.
    """
    out: dict[str, Any] = {}
    m = _MSG_DATE.search(msg)
    if m:
        out["listDate"] = m.group(1)
    m = _MSG_COUNTS.search(msg)
    if m:
        out["counts"] = {"black": int(m.group(1)), "grey": int(m.group(2))}
    m = _MSG_UPLOADER.search(msg)
    if m:
        out["uploader"], out["uploadedAt"] = m.group(1), m.group(2).strip()
    return out


@router.get("/history")
async def history(limit: int = 30) -> dict[str, Any]:
    """Every version of the baseline, newest first."""
    r = await _gh("GET", f"/repos/{GH_REPO}/commits",
                  params={"path": BASELINE_PATH, "sha": GH_BRANCH,
                          "per_page": max(1, min(limit, 100))})
    if r.status_code >= 400:
        raise HTTPException(502, f"读取版本历史失败：{r.status_code} {r.text[:180]}")
    items = []
    for c in r.json():
        msg = (c.get("commit") or {}).get("message") or ""
        items.append({
            "sha": c.get("sha"),
            "shortSha": (c.get("sha") or "")[:8],
            "committedAt": ((c.get("commit") or {}).get("committer") or {}).get("date"),
            "author": ((c.get("commit") or {}).get("author") or {}).get("name"),
            "subject": msg.split("\n")[0],
            "viaUploadPage": ("通过基线上传页提交" in msg
                              or "通过名单维护页提交" in msg),
            "commitUrl": c.get("html_url"),
            **_from_message(msg),
        })
    return {"repo": f"{GH_REPO}:{GH_BRANCH}", "path": BASELINE_PATH,
            "count": len(items), "versions": items}


@router.get("/history/{sha}")
async def history_detail(sha: str) -> dict[str, Any]:
    """The baseline exactly as it stood at one version, plus what that change did."""
    if not re.fullmatch(r"[0-9a-fA-F]{7,40}", sha):
        raise HTTPException(422, "版本号格式不正确")

    async def at(ref: str) -> dict[str, Any] | None:
        rr = await _gh("GET", f"/repos/{GH_REPO}/contents/{BASELINE_PATH}",
                       params={"ref": ref})
        if rr.status_code == 404:
            return None
        if rr.status_code >= 400:
            raise HTTPException(502, f"读取该版本失败：{rr.status_code} {rr.text[:160]}")
        return parse_workbook(base64.b64decode(rr.json()["content"]))

    async def json_at(path: str, ref: str) -> Optional[dict[str, Any]]:
        rr = await _gh("GET", f"/repos/{GH_REPO}/contents/{path}", params={"ref": ref})
        if rr.status_code == 404:
            return None
        if rr.status_code >= 400:
            return None
        try:
            return json.loads(base64.b64decode(rr.json()["content"]))
        except Exception:
            return None

    this = await at(sha)
    if not this:
        raise HTTPException(404, "该版本里没有基线文件")

    # The parent commit of this one gives "what changed in this version".
    prev = None
    rc = await _gh("GET", f"/repos/{GH_REPO}/commits/{sha}")
    if rc.status_code < 400:
        parents = rc.json().get("parents") or []
        if parents:
            try:
                prev = await at(parents[0]["sha"])
            except HTTPException:
                prev = None
        meta = rc.json()
        commit_info = {
            "sha": meta.get("sha"),
            "committedAt": ((meta.get("commit") or {}).get("committer") or {}).get("date"),
            "author": ((meta.get("commit") or {}).get("author") or {}).get("name"),
            "message": (meta.get("commit") or {}).get("message"),
            "commitUrl": meta.get("html_url"),
            **_from_message((meta.get("commit") or {}).get("message") or ""),
        }
    else:
        commit_info = {"sha": sha}

    # The other two lists as they stood at this version, and what this version did to
    # them. One upload is one commit, so the parent is the right thing to diff against.
    parent_sha = ((rc.json().get("parents") or [{}])[0].get("sha")
                  if rc.status_code < 400 else None)
    others: dict[str, Any] = {}
    for key, path, field in (("cpi", CPI_OVERRIDE_PATH, "scores"),
                             ("offshore", OFFSHORE_OVERRIDE_PATH, "jurisdictions")):
        cur_ov = await json_at(path, sha)
        if not cur_ov:
            others[key] = {"present": False}
            continue
        prev_ov = await json_at(path, parent_sha) if parent_sha else None
        vals = cur_ov.get(field) or {}
        names_now = set(vals.keys() if isinstance(vals, dict) else vals)
        entry: dict[str, Any] = {
            "present": True, "count": len(names_now),
            "confirmedBy": cur_ov.get("confirmedBy"), "confirmedAt": cur_ov.get("confirmedAt"),
            "edition": cur_ov.get("edition"),
            "names": sorted(names_now),
            "downloadUrl": f"https://raw.githubusercontent.com/{GH_REPO}/{sha}/{path}",
        }
        if prev_ov:
            pv = prev_ov.get(field) or {}
            names_before = set(pv.keys() if isinstance(pv, dict) else pv)
            entry["added"] = sorted(names_now - names_before)
            entry["removed"] = sorted(names_before - names_now)
            if isinstance(vals, dict) and isinstance(pv, dict):
                ch = [{"name": k, "from": pv[k], "to": vals[k]}
                      for k in sorted(names_now & names_before) if pv[k] != vals[k]]
                entry["valueChanged"] = ch[:60]
                entry["valueChangedCount"] = len(ch)
            entry["newInThisVersion"] = False
        else:
            entry["newInThisVersion"] = True
        others[key] = entry

    return {
        "commit": commit_info,
        "listDate": this["listDate"],
        "counts": {"black": len(this["normBlack"]), "grey": len(this["normGrey"])},
        "black": this["black"], "grey": this["grey"],
        "changeInThisVersion": diff(prev, this) if prev else None,
        "others": others,
        "downloadUrl": (f"https://raw.githubusercontent.com/{GH_REPO}/{sha}/"
                        f"{BASELINE_PATH}"),
    }


@router.get("/current")
async def current() -> dict[str, Any]:
    """What the pipeline is using right now, so the page can show it side by side."""
    try:
        cur_bytes, _ = await fetch_current()
    except HTTPException as e:
        return {"available": False, "error": e.detail}
    if not cur_bytes:
        return {"available": False, "error": "仓库里还没有基线文件"}
    cur = parse_workbook(cur_bytes)
    return {"available": True, "listDate": cur["listDate"],
            "counts": {"black": len(cur["normBlack"]), "grey": len(cur["normGrey"])},
            "black": cur["black"], "grey": cur["grey"],
            "repo": f"{GH_REPO}:{GH_BRANCH}", "path": BASELINE_PATH}
